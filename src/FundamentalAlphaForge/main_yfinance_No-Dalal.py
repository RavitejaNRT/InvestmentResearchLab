"""
FundamentalAlphaForge
=====================

Quantitative Equity Research Engine
===================================

CURRENT STAGE
-------------

Market Data + Fundamental Research Layer.

The engine:

1. Refreshes the current Nifty 500 universe.
2. Loads universe.py.
3. Downloads 2 years of daily OHLCV data.
4. Calculates market factors.
5. Calculates market scores.
6. Downloads fundamental information from Yahoo Finance via yfinance.
7. Calculates Quality, Growth and Valuation scores.
8. Calculates fundamental data completeness.
9. Calculates fundamental confidence.
10. Calculates Combined Research Score.
11. Produces research rankings and diagnostics.
12. Saves the complete research dataset to Excel.
13. Builds an Excel Dashboard with visual statistics.
14. Reports complete program runtime with timestamps.

IMPORTANT FUNDAMENTAL METHODOLOGY
---------------------------------

The fundamental model deliberately does NOT treat missing values as zero.

Missing values remain NaN.

The model uses available information while separately tracking
data completeness and confidence.

QUALITY
-------

Current Quality factors:

    ROE
    ROA
    Debt / Equity
    Profit Margin
    Operating Margin
    Gross Margin
    Current Ratio
    Quick Ratio
    Free Cash Flow

GROWTH
------

Current Growth factors:

    Revenue Growth
    Earnings Growth
    Quarterly Revenue Growth

IMPORTANT:

Yahoo's earningsQuarterlyGrowth is NOT treated as EPS Growth.

VALUATION
---------

Current Valuation factors:

    P/E
    Forward P/E
    P/B
    PEG
    Price / Sales
    EV / EBITDA

DATA COVERAGE
-------------

The engine reports actual coverage for every fundamental factor.

CONFIDENCE
----------

High:
    >= 80% fundamental completeness

Medium:
    >= 60% and < 80%

Low:
    < 60%

Low-confidence stocks are excluded from the displayed Fundamental
and Combined headline rankings.

They remain in the Excel Research Data sheet.

CURRENT RESEARCH VS BACKTESTING
-------------------------------

Current Yahoo Finance fundamental information represents currently
available/latest information.

It is NOT a complete point-in-time historical fundamental database.

Therefore:

    CURRENT RESEARCH = YES

    HISTORICAL FUNDAMENTAL BACKTEST = NOT YET VALID

Before using fundamentals inside historical VectorBT testing,
financial statement reporting/availability dates must be handled
to prevent look-ahead bias.
"""

from __future__ import annotations

import importlib.util
import io
import sys
import time
from contextlib import redirect_stdout
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    ScatterChart,
    Reference,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


from trade_data import (
    refresh_nifty500_universe,
    get_historical_market_data_for_symbols,
)


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

UNIVERSE_FILE = PROJECT_ROOT / "universe.py"

RESULTS_DIR = PROJECT_ROOT / "results"


# ============================================================
# MARKET DATA PARAMETERS
# ============================================================

HISTORICAL_PERIOD = "2y"

MIN_PRICE = 100.0

MIN_TRADING_DAYS = 200

MOMENTUM_3M_DAYS = 63

MOMENTUM_6M_DAYS = 126

MOMENTUM_12M_DAYS = 252

SHORT_MA = 50

LONG_MA = 200

VOLATILITY_DAYS = 63

AVERAGE_VOLUME_DAYS = 20


# ============================================================
# FUNDAMENTAL PARAMETERS
# ============================================================

FUNDAMENTAL_SLEEP_SECONDS = 0.05

FUNDAMENTAL_PROGRESS_INTERVAL = 25


# ============================================================
# FUNDAMENTAL CONFIDENCE
# ============================================================

HIGH_CONFIDENCE_COMPLETENESS = 80.0

MEDIUM_CONFIDENCE_COMPLETENESS = 60.0


# ============================================================
# FUNDAMENTAL FACTORS
# ============================================================

ALL_FUNDAMENTAL_FACTORS = [

    # Quality
    "roe",
    "roa",
    "debt_equity",
    "profit_margin",
    "operating_margin",
    "gross_margin",
    "current_ratio",
    "quick_ratio",
    "free_cash_flow",

    # Growth
    "revenue_growth",
    "earnings_growth",
    "quarterly_revenue_growth",

    # Valuation
    "pe",
    "forward_pe",
    "price_book",
    "peg",
    "price_sales",
    "ev_ebitda",
]


TOTAL_FUNDAMENTAL_FACTORS = len(
    ALL_FUNDAMENTAL_FACTORS
)


# ============================================================
# FACTOR GROUPS
# ============================================================

FUNDAMENTAL_FACTOR_GROUPS = {

    "quality": [
        "roe",
        "roa",
        "debt_equity",
        "profit_margin",
        "operating_margin",
        "gross_margin",
        "current_ratio",
        "quick_ratio",
        "free_cash_flow",
    ],

    "growth": [
        "revenue_growth",
        "earnings_growth",
        "quarterly_revenue_growth",
    ],

    "valuation": [
        "pe",
        "forward_pe",
        "price_book",
        "peg",
        "price_sales",
        "ev_ebitda",
    ],
}


# ============================================================
# FUNDAMENTAL FACTOR WEIGHTS
# ============================================================

QUALITY_FACTOR_WEIGHTS = {

    "roe": 0.20,
    "roa": 0.10,
    "debt_equity": 0.15,
    "profit_margin": 0.10,
    "operating_margin": 0.10,
    "gross_margin": 0.05,
    "current_ratio": 0.05,
    "quick_ratio": 0.05,
    "free_cash_flow": 0.20,
}


GROWTH_FACTOR_WEIGHTS = {

    "revenue_growth": 0.40,
    "earnings_growth": 0.40,
    "quarterly_revenue_growth": 0.20,
}


VALUATION_FACTOR_WEIGHTS = {

    "pe": 0.20,
    "forward_pe": 0.15,
    "price_book": 0.10,
    "peg": 0.10,
    "price_sales": 0.20,
    "ev_ebitda": 0.25,
}


# ============================================================
# SCORE WEIGHTS
# ============================================================

QUALITY_WEIGHTS = {
    f"score_{factor}": weight
    for factor, weight
    in QUALITY_FACTOR_WEIGHTS.items()
}


GROWTH_WEIGHTS = {
    f"score_{factor}": weight
    for factor, weight
    in GROWTH_FACTOR_WEIGHTS.items()
}


VALUATION_WEIGHTS = {

    "score_pe":
        VALUATION_FACTOR_WEIGHTS["pe"],

    "score_forward_pe":
        VALUATION_FACTOR_WEIGHTS["forward_pe"],

    "score_pb":
        VALUATION_FACTOR_WEIGHTS["price_book"],

    "score_peg":
        VALUATION_FACTOR_WEIGHTS["peg"],

    "score_price_sales":
        VALUATION_FACTOR_WEIGHTS["price_sales"],

    "score_ev_ebitda":
        VALUATION_FACTOR_WEIGHTS["ev_ebitda"],
}


# ============================================================
# FUNDAMENTAL GROUP WEIGHTS
# ============================================================

FUNDAMENTAL_WEIGHTS = {

    "quality_score": 0.35,

    "growth_score": 0.35,

    "valuation_score": 0.30,
}


# ============================================================
# COMBINED RESEARCH SCORE
# ============================================================

COMBINED_RESEARCH_WEIGHTS = {

    "market_research_score": 0.50,

    "fundamental_score": 0.50,
}


# ============================================================
# DISPLAY PARAMETERS
# ============================================================

TOP_STOCKS_TO_DISPLAY = 30

TOP_STOCKS_DETAILED = 10

TOP_FACTOR_LEADERS = 5

TOP_RESEARCH_CANDIDATES = 10


# ============================================================
# MARKET SCORE WEIGHTS
# ============================================================

MOMENTUM_WEIGHTS = {

    "score_3m": 0.20,

    "score_6m": 0.20,

    "score_12m": 0.20,

    "score_price_ma50": 0.10,

    "score_price_ma200": 0.10,

    "score_ma50_ma200": 0.10,

    "score_52w_proximity": 0.10,
}


MARKET_RESEARCH_WEIGHTS = {

    "momentum_score": 0.50,

    "trend_score": 0.30,

    "risk_score": 0.20,
}


# ============================================================
# EXCEL CONFIGURATION
# ============================================================

EXCEL_OUTPUT_FILE = (
    RESULTS_DIR
    / "fundamental_alpha_forge_results.xlsx"
)

RESEARCH_DATA_SHEET = "Research Data"

DASHBOARD_SHEET = "Dashboard"

COVERAGE_SHEET = "Factor Coverage"

STATISTICS_SHEET = "Score Statistics"


# ============================================================
# RUNTIME TRACKING
# ============================================================

PROGRAM_START_TIME = None

PROGRAM_END_TIME = None

PROGRAM_ELAPSED_SECONDS = None


# ============================================================
# PRINT HELPERS
# ============================================================

def print_header(title: str) -> None:

    print()
    print("=" * 80)
    print(title)
    print("=" * 80)


def print_subheader(title: str) -> None:

    print()
    print("-" * 80)
    print(title)
    print("-" * 80)


def safe_float(value) -> float:

    try:

        if pd.isna(value):
            return np.nan

        return float(value)

    except (TypeError, ValueError):

        return np.nan


# ============================================================
# DISPLAY FORMATTING
# ============================================================

def format_display_table(
    table: pd.DataFrame,
) -> pd.DataFrame:

    result = table.copy()

    for column in result.columns:

        if pd.api.types.is_numeric_dtype(
            result[column]
        ):

            result[column] = (
                result[column]
                .round(2)
            )

    result = result.replace(
        {
            np.nan: "N/A",
            np.inf: "N/A",
            -np.inf: "N/A",
        }
    )

    return result


# ============================================================
# UNIVERSE
# ============================================================

def load_universe_from_file() -> list[str]:

    if not UNIVERSE_FILE.exists():

        raise FileNotFoundError(
            f"Universe file not found:\n{UNIVERSE_FILE}"
        )

    spec = importlib.util.spec_from_file_location(
        "generated_universe",
        UNIVERSE_FILE,
    )

    if spec is None or spec.loader is None:

        raise ImportError(
            "Could not load universe.py."
        )

    module = importlib.util.module_from_spec(
        spec
    )

    spec.loader.exec_module(module)

    symbols = getattr(
        module,
        "symbols",
        None,
    )

    if symbols is None:

        raise AttributeError(
            "universe.py does not contain 'symbols'."
        )

    if not isinstance(symbols, list):

        raise TypeError(
            "universe.py 'symbols' must be a list."
        )

    symbols = [
        str(symbol).strip().upper()
        for symbol in symbols
        if str(symbol).strip()
    ]

    symbols = sorted(
        set(symbols)
    )

    if len(symbols) < 400:

        raise ValueError(
            f"Only {len(symbols)} symbols found in universe.py. "
            "Expected a valid Nifty 500 universe."
        )

    return symbols


def refresh_and_load_universe() -> list[str]:

    print_header(
        "FUNDAMENTALALPHAFORGE — UNIVERSE"
    )

    print(
        f"Project root : {PROJECT_ROOT}"
    )

    print(
        f"Universe file: {UNIVERSE_FILE}"
    )

    print()

    print(
        "Refreshing current Nifty 500 constituents..."
    )

    refresh_nifty500_universe()

    symbols = load_universe_from_file()

    print()

    print(
        f"Loaded universe: {len(symbols):,} symbols"
    )

    return symbols


# ============================================================
# MARKET DATA EXTRACTION
# ============================================================

def extract_symbol_series(
    data: pd.DataFrame,
    field: str,
    symbol: str,
) -> Optional[pd.Series]:

    try:

        if data is None or data.empty:
            return None

        if field not in data.columns:
            return None

        series = data[field]

        if isinstance(
            series,
            pd.DataFrame,
        ):

            if symbol not in series.columns:
                return None

            series = series[symbol]

        elif isinstance(
            series,
            pd.Series,
        ):

            pass

        else:

            return None

        series = pd.to_numeric(
            series,
            errors="coerce",
        )

        series = series.replace(
            [
                np.inf,
                -np.inf,
            ],
            np.nan,
        )

        series = series.dropna()

        if series.empty:
            return None

        return series

    except (
        KeyError,
        TypeError,
        AttributeError,
    ):

        return None


# ============================================================
# RETURNS
# ============================================================

def calculate_returns(
    close: pd.Series,
) -> dict:

    result = {

        "return_3m": np.nan,

        "return_6m": np.nan,

        "return_12m": np.nan,
    }

    if close.empty:
        return result

    current_price = safe_float(
        close.iloc[-1]
    )

    if (
        pd.isna(current_price)
        or current_price <= 0
    ):
        return result

    if len(close) > MOMENTUM_3M_DAYS:

        previous_price = safe_float(
            close.iloc[
                -MOMENTUM_3M_DAYS - 1
            ]
        )

        if (
            pd.notna(previous_price)
            and previous_price > 0
        ):

            result["return_3m"] = (
                current_price
                / previous_price
                - 1
            ) * 100

    if len(close) > MOMENTUM_6M_DAYS:

        previous_price = safe_float(
            close.iloc[
                -MOMENTUM_6M_DAYS - 1
            ]
        )

        if (
            pd.notna(previous_price)
            and previous_price > 0
        ):

            result["return_6m"] = (
                current_price
                / previous_price
                - 1
            ) * 100

    if len(close) > MOMENTUM_12M_DAYS:

        previous_price = safe_float(
            close.iloc[
                -MOMENTUM_12M_DAYS - 1
            ]
        )

        if (
            pd.notna(previous_price)
            and previous_price > 0
        ):

            result["return_12m"] = (
                current_price
                / previous_price
                - 1
            ) * 100

    return result


# ============================================================
# MOVING AVERAGES
# ============================================================

def calculate_moving_averages(
    close: pd.Series,
) -> dict:

    result = {

        "ma50": np.nan,

        "ma200": np.nan,

        "price_vs_ma50": np.nan,

        "price_vs_ma200": np.nan,

        "ma50_vs_ma200": np.nan,
    }

    if close.empty:
        return result

    current_price = safe_float(
        close.iloc[-1]
    )

    if len(close) >= SHORT_MA:

        result["ma50"] = safe_float(
            close
            .rolling(SHORT_MA)
            .mean()
            .iloc[-1]
        )

    if len(close) >= LONG_MA:

        result["ma200"] = safe_float(
            close
            .rolling(LONG_MA)
            .mean()
            .iloc[-1]
        )

    if (
        pd.notna(result["ma50"])
        and result["ma50"] > 0
    ):

        result["price_vs_ma50"] = (
            current_price
            / result["ma50"]
            - 1
        ) * 100

    if (
        pd.notna(result["ma200"])
        and result["ma200"] > 0
    ):

        result["price_vs_ma200"] = (
            current_price
            / result["ma200"]
            - 1
        ) * 100

    if (
        pd.notna(result["ma50"])
        and pd.notna(result["ma200"])
        and result["ma200"] > 0
    ):

        result["ma50_vs_ma200"] = (
            result["ma50"]
            / result["ma200"]
            - 1
        ) * 100

    return result


# ============================================================
# 52 WEEK METRICS
# ============================================================

def calculate_52_week_metrics(
    close: pd.Series,
) -> dict:

    result = {

        "high_52w": np.nan,

        "low_52w": np.nan,

        "distance_from_52w_high": np.nan,

        "distance_from_52w_low": np.nan,
    }

    if close.empty:
        return result

    window = min(
        MOMENTUM_12M_DAYS,
        len(close),
    )

    recent = close.tail(
        window
    )

    if recent.empty:
        return result

    current_price = safe_float(
        close.iloc[-1]
    )

    high_52w = safe_float(
        recent.max()
    )

    low_52w = safe_float(
        recent.min()
    )

    result["high_52w"] = high_52w
    result["low_52w"] = low_52w

    if (
        pd.notna(high_52w)
        and high_52w > 0
    ):

        result[
            "distance_from_52w_high"
        ] = (
            current_price
            / high_52w
            - 1
        ) * 100

    if (
        pd.notna(low_52w)
        and low_52w > 0
    ):

        result[
            "distance_from_52w_low"
        ] = (
            current_price
            / low_52w
            - 1
        ) * 100

    return result


# ============================================================
# VOLATILITY
# ============================================================

def calculate_volatility(
    close: pd.Series,
) -> float:

    if close.empty:
        return np.nan

    daily_returns = (
        close
        .pct_change()
        .replace(
            [
                np.inf,
                -np.inf,
            ],
            np.nan,
        )
        .dropna()
    )

    recent_returns = (
        daily_returns.tail(
            VOLATILITY_DAYS
        )
    )

    if len(recent_returns) < 20:
        return np.nan

    return safe_float(
        recent_returns.std()
        * np.sqrt(252)
        * 100
    )


# ============================================================
# MAXIMUM DRAWDOWN
# ============================================================

def calculate_max_drawdown(
    close: pd.Series,
) -> float:

    if close.empty:
        return np.nan

    running_peak = close.cummax()

    drawdown = (
        close
        / running_peak
        - 1
    ) * 100

    return safe_float(
        drawdown.min()
    )


# ============================================================
# VOLUME
# ============================================================

def calculate_volume_metrics(
    volume: Optional[pd.Series],
) -> dict:

    result = {

        "avg_volume_20d": np.nan,

        "current_volume": np.nan,

        "volume_ratio": np.nan,
    }

    if volume is None or volume.empty:
        return result

    volume = pd.to_numeric(
        volume,
        errors="coerce",
    )

    volume = volume.replace(
        [
            np.inf,
            -np.inf,
        ],
        np.nan,
    ).dropna()

    if volume.empty:
        return result

    recent = volume.tail(
        AVERAGE_VOLUME_DAYS
    )

    avg_volume = safe_float(
        recent.mean()
    )

    current_volume = safe_float(
        volume.iloc[-1]
    )

    result[
        "avg_volume_20d"
    ] = avg_volume

    result[
        "current_volume"
    ] = current_volume

    if (
        pd.notna(avg_volume)
        and avg_volume > 0
    ):

        result[
            "volume_ratio"
        ] = (
            current_volume
            / avg_volume
        )

    return result


# ============================================================
# STOCK MARKET METRICS
# ============================================================

def calculate_stock_metrics(
    data: pd.DataFrame,
    symbol: str,
) -> Optional[dict]:

    close = extract_symbol_series(
        data,
        "Close",
        symbol,
    )

    if close is None:
        return None

    if len(close) < MIN_TRADING_DAYS:
        return None

    current_price = safe_float(
        close.iloc[-1]
    )

    if (
        pd.isna(current_price)
        or current_price < MIN_PRICE
    ):
        return None

    volume = extract_symbol_series(
        data,
        "Volume",
        symbol,
    )

    result = {

        "symbol": symbol,

        "price": current_price,

        "data_points": len(close),
    }

    result.update(
        calculate_returns(close)
    )

    result.update(
        calculate_moving_averages(close)
    )

    result.update(
        calculate_52_week_metrics(close)
    )

    result["volatility"] = (
        calculate_volatility(close)
    )

    result["max_drawdown"] = (
        calculate_max_drawdown(close)
    )

    result.update(
        calculate_volume_metrics(volume)
    )

    result["above_ma50"] = (
        pd.notna(result["ma50"])
        and current_price
        > result["ma50"]
    )

    result["above_ma200"] = (
        pd.notna(result["ma200"])
        and current_price
        > result["ma200"]
    )

    result["ma50_above_ma200"] = (
        pd.notna(result["ma50"])
        and pd.notna(result["ma200"])
        and result["ma50"]
        > result["ma200"]
    )

    result["strong_uptrend"] = (
        result["above_ma50"]
        and result["above_ma200"]
        and result["ma50_above_ma200"]
    )

    return result


# ============================================================
# BUILD MARKET DATAFRAME
# ============================================================

def build_research_dataframe(
    data: pd.DataFrame,
    valid_symbols: list[str],
) -> pd.DataFrame:

    print_header(
        "CALCULATING MARKET FACTORS"
    )

    records = []

    total = len(valid_symbols)

    for index, symbol in enumerate(
        valid_symbols,
        start=1,
    ):

        metrics = calculate_stock_metrics(
            data,
            symbol,
        )

        if metrics is not None:
            records.append(metrics)

        if (
            index % 50 == 0
            or index == total
        ):

            print(
                f"Processed {index:,} / "
                f"{total:,} symbols"
            )

    if not records:

        raise RuntimeError(
            "No stocks passed the market-data processing stage."
        )

    df = pd.DataFrame(
        records
    )

    print()

    print(
        f"Stocks with usable data: {len(df):,}"
    )

    return df


# ============================================================
# PERCENTILE SCORE
# ============================================================

def percentile_score(
    series: pd.Series,
    higher_is_better: bool = True,
) -> pd.Series:

    numeric = pd.to_numeric(
        series,
        errors="coerce",
    )

    score = (
        numeric
        .rank(
            pct=True,
            method="average",
        )
        * 100
    )

    if not higher_is_better:
        score = 100 - score

    return score


# ============================================================
# WEIGHTED SCORE
# ============================================================

def weighted_score(
    df: pd.DataFrame,
    weights: dict[str, float],
) -> pd.Series:

    numerator = pd.Series(
        0.0,
        index=df.index,
    )

    denominator = pd.Series(
        0.0,
        index=df.index,
    )

    for column, weight in weights.items():

        if column not in df.columns:
            continue

        values = pd.to_numeric(
            df[column],
            errors="coerce",
        )

        valid = values.notna()

        numerator.loc[valid] += (
            values.loc[valid]
            * weight
        )

        denominator.loc[valid] += weight

    result = (
        numerator
        / denominator
    )

    result = result.where(
        denominator > 0
    )

    return result


# ============================================================
# WEIGHTED COMPLETENESS
# ============================================================

def weighted_completeness(
    df: pd.DataFrame,
    columns_and_weights: dict[str, float],
) -> pd.Series:

    completeness = pd.Series(
        0.0,
        index=df.index,
    )

    total_weight = sum(
        columns_and_weights.values()
    )

    if total_weight <= 0:

        return pd.Series(
            np.nan,
            index=df.index,
        )

    for column, weight in (
        columns_and_weights.items()
    ):

        if column not in df.columns:
            continue

        valid = df[column].notna()

        completeness.loc[valid] += weight

    completeness = (
        completeness
        / total_weight
        * 100
    )

    return completeness.round(2)


# ============================================================
# SIMPLE COMPLETENESS
# ============================================================

def calculate_simple_completeness(
    df: pd.DataFrame,
    columns: list[str],
) -> pd.Series:

    available_columns = [
        column
        for column in columns
        if column in df.columns
    ]

    if not available_columns:

        return pd.Series(
            np.nan,
            index=df.index,
        )

    return (
        df[available_columns]
        .notna()
        .sum(axis=1)
        / len(available_columns)
        * 100
    ).round(2)


# ============================================================
# CONFIDENCE
# ============================================================

def fundamental_confidence_label(
    completeness: float,
) -> str:

    completeness = safe_float(
        completeness
    )

    if pd.isna(completeness):
        return "No Data"

    if completeness >= HIGH_CONFIDENCE_COMPLETENESS:
        return "High"

    if completeness >= MEDIUM_CONFIDENCE_COMPLETENESS:
        return "Medium"

    return "Low"


# ============================================================
# MARKET SCORES
# ============================================================

def calculate_momentum_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df["score_3m"] = percentile_score(
        df["return_3m"]
    )

    df["score_6m"] = percentile_score(
        df["return_6m"]
    )

    df["score_12m"] = percentile_score(
        df["return_12m"]
    )

    df["score_price_ma50"] = percentile_score(
        df["price_vs_ma50"]
    )

    df["score_price_ma200"] = percentile_score(
        df["price_vs_ma200"]
    )

    df["score_ma50_ma200"] = percentile_score(
        df["ma50_vs_ma200"]
    )

    df["score_52w_proximity"] = percentile_score(
        df["distance_from_52w_high"]
    )

    df["momentum_score"] = (
        weighted_score(
            df,
            MOMENTUM_WEIGHTS,
        )
        .round(2)
    )

    return df


def calculate_trend_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    trend_columns = [
        "above_ma50",
        "above_ma200",
        "ma50_above_ma200",
    ]

    available = (
        df[trend_columns]
        .notna()
        .sum(axis=1)
    )

    positive = (
        df[trend_columns]
        .fillna(False)
        .astype(int)
        .sum(axis=1)
    )

    df["trend_score"] = (
        positive
        / available.replace(
            0,
            np.nan,
        )
        * 100
    ).round(2)

    return df


def calculate_risk_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    volatility_score = percentile_score(
        df["volatility"],
        higher_is_better=False,
    )

    drawdown_score = percentile_score(
        df["max_drawdown"],
        higher_is_better=True,
    )

    risk_df = pd.DataFrame(
        {
            "volatility_score":
                volatility_score,

            "drawdown_score":
                drawdown_score,
        },
        index=df.index,
    )

    df["risk_score"] = (
        weighted_score(
            risk_df,
            {
                "volatility_score": 0.50,
                "drawdown_score": 0.50,
            },
        )
        .round(2)
    )

    return df


def calculate_market_research_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df["market_research_score"] = (
        weighted_score(
            df,
            MARKET_RESEARCH_WEIGHTS,
        )
        .round(2)
    )

    return df


# ============================================================
# FUNDAMENTAL HELPERS
# ============================================================

def get_info_value(
    info: dict,
    keys: list[str],
) -> float:

    for key in keys:

        value = info.get(key)

        value = safe_float(value)

        if pd.notna(value):
            return value

    return np.nan


def normalize_percentage(
    value: float,
) -> float:

    value = safe_float(value)

    if pd.isna(value):
        return np.nan

    if abs(value) <= 1:
        return value * 100

    return value


def normalize_ratio(
    value: float,
) -> float:

    value = safe_float(value)

    if pd.isna(value):
        return np.nan

    return value


# ============================================================
# FUNDAMENTAL DATA
# ============================================================

def fetch_fundamental_data(
    symbol: str,
) -> Optional[dict]:

    """
    Fetch current/latest fundamental information through yfinance.

    Missing values are NOT converted to zero.

    EPS growth is deliberately NOT fetched from
    earningsQuarterlyGrowth.
    """

    try:

        ticker = yf.Ticker(symbol)

        info = ticker.info

        if not isinstance(info, dict):
            return None

        # ====================================================
        # QUALITY
        # ====================================================

        roe = get_info_value(
            info,
            ["returnOnEquity"],
        )

        roa = get_info_value(
            info,
            ["returnOnAssets"],
        )

        debt_equity = get_info_value(
            info,
            ["debtToEquity"],
        )

        profit_margin = get_info_value(
            info,
            ["profitMargins"],
        )

        operating_margin = get_info_value(
            info,
            ["operatingMargins"],
        )

        gross_margin = get_info_value(
            info,
            ["grossMargins"],
        )

        current_ratio = get_info_value(
            info,
            ["currentRatio"],
        )

        quick_ratio = get_info_value(
            info,
            ["quickRatio"],
        )

        free_cash_flow = get_info_value(
            info,
            ["freeCashflow"],
        )

        if (
            pd.notna(debt_equity)
            and debt_equity > 10
        ):

            debt_equity = (
                debt_equity / 100
            )

        # ====================================================
        # GROWTH
        # ====================================================

        revenue_growth = get_info_value(
            info,
            ["revenueGrowth"],
        )

        earnings_growth = get_info_value(
            info,
            ["earningsGrowth"],
        )

        quarterly_revenue_growth = get_info_value(
            info,
            ["revenueQuarterlyGrowth"],
        )

        revenue_growth = normalize_percentage(
            revenue_growth
        )

        earnings_growth = normalize_percentage(
            earnings_growth
        )

        quarterly_revenue_growth = (
            normalize_percentage(
                quarterly_revenue_growth
            )
        )

        # ====================================================
        # VALUATION
        # ====================================================

        pe = get_info_value(
            info,
            ["trailingPE"],
        )

        forward_pe = get_info_value(
            info,
            ["forwardPE"],
        )

        price_book = get_info_value(
            info,
            ["priceToBook"],
        )

        peg = get_info_value(
            info,
            ["pegRatio"],
        )

        price_sales = get_info_value(
            info,
            ["priceToSalesTrailing12Months"],
        )

        ev_ebitda = get_info_value(
            info,
            ["enterpriseToEbitda"],
        )

        # ====================================================
        # ADDITIONAL INFORMATION
        # ====================================================

        market_cap = get_info_value(
            info,
            ["marketCap"],
        )

        enterprise_value = get_info_value(
            info,
            ["enterpriseValue"],
        )

        return {

            "symbol": symbol,

            # Quality
            "roe": (
                roe * 100
                if pd.notna(roe)
                else np.nan
            ),

            "roa": (
                roa * 100
                if pd.notna(roa)
                else np.nan
            ),

            "debt_equity":
                debt_equity,

            "profit_margin": (
                profit_margin * 100
                if pd.notna(profit_margin)
                else np.nan
            ),

            "operating_margin": (
                operating_margin * 100
                if pd.notna(operating_margin)
                else np.nan
            ),

            "gross_margin": (
                gross_margin * 100
                if pd.notna(gross_margin)
                else np.nan
            ),

            "current_ratio":
                current_ratio,

            "quick_ratio":
                quick_ratio,

            "free_cash_flow":
                free_cash_flow,

            # Growth
            "revenue_growth":
                revenue_growth,

            "earnings_growth":
                earnings_growth,

            "quarterly_revenue_growth":
                quarterly_revenue_growth,

            # Valuation
            "pe":
                pe,

            "forward_pe":
                forward_pe,

            "price_book":
                price_book,

            "peg":
                peg,

            "price_sales":
                price_sales,

            "ev_ebitda":
                ev_ebitda,

            # Additional
            "market_cap":
                market_cap,

            "enterprise_value":
                enterprise_value,
        }

    except Exception:

        return None


# ============================================================
# FUNDAMENTAL COLLECTION
# ============================================================

def build_fundamental_dataframe(
    symbols: list[str],
) -> pd.DataFrame:

    print_header(
        "DOWNLOADING FUNDAMENTAL DATA"
    )

    print(
        "Source : Yahoo Finance via yfinance"
    )

    print()

    print(
        "Expanded fundamental model:"
    )

    print(
        "Quality : ROE, ROA, D/E, "
        "Profit Margin, Operating Margin, "
        "Gross Margin, Current Ratio, "
        "Quick Ratio, Free Cash Flow"
    )

    print(
        "Growth  : Revenue Growth, "
        "Earnings Growth, "
        "Quarterly Revenue Growth"
    )

    print(
        "Value   : P/E, Forward P/E, P/B, "
        "PEG, Price/Sales, EV/EBITDA"
    )

    print()

    print(
        "EPS Growth proxy has been REMOVED."
    )

    print(
        "Missing fields remain NaN."
    )

    print()

    records = []

    total = len(symbols)

    for index, symbol in enumerate(
        symbols,
        start=1,
    ):

        fundamentals = (
            fetch_fundamental_data(symbol)
        )

        if fundamentals is not None:

            records.append(
                fundamentals
            )

        if (
            index % FUNDAMENTAL_PROGRESS_INTERVAL == 0
            or index == total
        ):

            print(
                f"Fundamentals processed "
                f"{index:,} / {total:,}"
            )

        if FUNDAMENTAL_SLEEP_SECONDS > 0:

            time.sleep(
                FUNDAMENTAL_SLEEP_SECONDS
            )

    if not records:

        print(
            "WARNING: No fundamental data was retrieved."
        )

        return pd.DataFrame(
            columns=[

                "symbol",

                "roe",
                "roa",
                "debt_equity",
                "profit_margin",
                "operating_margin",
                "gross_margin",
                "current_ratio",
                "quick_ratio",
                "free_cash_flow",

                "revenue_growth",
                "earnings_growth",
                "quarterly_revenue_growth",

                "pe",
                "forward_pe",
                "price_book",
                "peg",
                "price_sales",
                "ev_ebitda",

                "market_cap",
                "enterprise_value",
            ]
        )

    fundamentals_df = pd.DataFrame(
        records
    )

    print()

    print(
        f"Stocks with fundamental data: "
        f"{len(fundamentals_df):,}"
    )

    return fundamentals_df


# ============================================================
# FUNDAMENTAL DATA QUALITY
# ============================================================

def calculate_fundamental_data_quality(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    available_factor_columns = [
        column
        for column in ALL_FUNDAMENTAL_FACTORS
        if column in df.columns
    ]

    if available_factor_columns:

        df[
            "fundamental_factors_available"
        ] = (
            df[
                available_factor_columns
            ]
            .notna()
            .sum(axis=1)
        )

        df[
            "fundamental_simple_completeness"
        ] = (
            df[
                "fundamental_factors_available"
            ]
            / TOTAL_FUNDAMENTAL_FACTORS
            * 100
        ).round(2)

    else:

        df[
            "fundamental_factors_available"
        ] = 0

        df[
            "fundamental_simple_completeness"
        ] = 0.0

    # ========================================================
    # SIMPLE GROUP COMPLETENESS
    # ========================================================

    df[
        "quality_data_completeness"
    ] = calculate_simple_completeness(
        df,
        FUNDAMENTAL_FACTOR_GROUPS["quality"],
    )

    df[
        "growth_data_completeness"
    ] = calculate_simple_completeness(
        df,
        FUNDAMENTAL_FACTOR_GROUPS["growth"],
    )

    df[
        "valuation_data_completeness"
    ] = calculate_simple_completeness(
        df,
        FUNDAMENTAL_FACTOR_GROUPS["valuation"],
    )

    # ========================================================
    # WEIGHTED GROUP COMPLETENESS
    # ========================================================

    df[
        "quality_weighted_completeness"
    ] = weighted_completeness(
        df,
        QUALITY_FACTOR_WEIGHTS,
    )

    df[
        "growth_weighted_completeness"
    ] = weighted_completeness(
        df,
        GROWTH_FACTOR_WEIGHTS,
    )

    df[
        "valuation_weighted_completeness"
    ] = weighted_completeness(
        df,
        VALUATION_FACTOR_WEIGHTS,
    )

    # ========================================================
    # OVERALL WEIGHTED COMPLETENESS
    # ========================================================

    df[
        "fundamental_weighted_completeness"
    ] = (
        df[
            "quality_weighted_completeness"
        ]
        * FUNDAMENTAL_WEIGHTS["quality_score"]

        +

        df[
            "growth_weighted_completeness"
        ]
        * FUNDAMENTAL_WEIGHTS["growth_score"]

        +

        df[
            "valuation_weighted_completeness"
        ]
        * FUNDAMENTAL_WEIGHTS["valuation_score"]
    ).round(2)

    df[
        "fundamental_data_completeness"
    ] = (
        df[
            "fundamental_weighted_completeness"
        ]
    )

    # ========================================================
    # CONFIDENCE
    # ========================================================

    df[
        "fundamental_confidence"
    ] = (
        df[
            "fundamental_data_completeness"
        ]
        .apply(
            fundamental_confidence_label
        )
    )

    # ========================================================
    # RANKING ELIGIBILITY
    # ========================================================

    df[
        "fundamental_ranking_eligible"
    ] = (
        df[
            "fundamental_confidence"
        ]
        .isin(
            [
                "High",
                "Medium",
            ]
        )
    )

    return df


# ============================================================
# FUNDAMENTAL AVAILABILITY
# ============================================================

def display_fundamental_availability(
    df: pd.DataFrame,
) -> None:

    print_header(
        "FUNDAMENTAL DATA AVAILABILITY"
    )

    rows = []

    total = len(df)

    for column in ALL_FUNDAMENTAL_FACTORS:

        if column not in df.columns:
            continue

        available = int(
            df[column].notna().sum()
        )

        missing = total - available

        coverage = (
            available / total * 100
            if total > 0
            else 0
        )

        rows.append(
            {
                "Factor": column,
                "Available": available,
                "Missing": missing,
                "Coverage %": round(
                    coverage,
                    2,
                ),
                ">80%": (
                    "YES"
                    if coverage >= 80
                    else "NO"
                ),
            }
        )

    if rows:

        table = pd.DataFrame(rows)

        table = table.sort_values(
            "Coverage %",
            ascending=False,
        )

        print(
            format_display_table(
                table
            ).to_string(
                index=False
            )
        )

    print_subheader(
        "GROUP-LEVEL DATA COMPLETENESS"
    )

    group_rows = []

    groups = {

        "Quality":
            "quality_weighted_completeness",

        "Growth":
            "growth_weighted_completeness",

        "Valuation":
            "valuation_weighted_completeness",

        "Overall":
            "fundamental_weighted_completeness",
    }

    for name, column in groups.items():

        if column not in df.columns:
            continue

        group_rows.append(
            {
                "Group": name,

                "Average Completeness %":
                    round(
                        df[column].mean(),
                        2,
                    ),

                "Median Completeness %":
                    round(
                        df[column].median(),
                        2,
                    ),

                ">=80%":
                    int(
                        (
                            df[column] >= 80
                        ).sum()
                    ),

                ">=60%":
                    int(
                        (
                            df[column] >= 60
                        ).sum()
                    ),
            }
        )

    if group_rows:

        print(
            pd.DataFrame(
                group_rows
            ).to_string(
                index=False
            )
        )

    print_subheader(
        "FUNDAMENTAL CONFIDENCE SUMMARY"
    )

    completeness = df[
        "fundamental_data_completeness"
    ]

    confidence_summary = pd.DataFrame(
        {
            "Metric": [

                "Average completeness",

                "Median completeness",

                "High confidence (>=80%)",

                "Medium confidence (60-79.99%)",

                "Low confidence (<60%)",

                "Eligible for ranking",

                "Excluded from headline ranking",
            ],

            "Value": [

                round(
                    completeness.mean(),
                    2,
                ),

                round(
                    completeness.median(),
                    2,
                ),

                int(
                    (
                        completeness >= 80
                    ).sum()
                ),

                int(
                    (
                        (completeness >= 60)
                        &
                        (completeness < 80)
                    ).sum()
                ),

                int(
                    (
                        completeness < 60
                    ).sum()
                ),

                int(
                    df[
                        "fundamental_ranking_eligible"
                    ].sum()
                ),

                int(
                    (
                        ~df[
                            "fundamental_ranking_eligible"
                        ]
                    ).sum()
                ),
            ],
        }
    )

    print(
        confidence_summary.to_string(
            index=False
        )
    )


# ============================================================
# QUALITY SCORE
# ============================================================

def calculate_quality_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df["score_roe"] = percentile_score(
        df["roe"],
        higher_is_better=True,
    )

    df["score_roa"] = percentile_score(
        df["roa"],
        higher_is_better=True,
    )

    df["score_debt_equity"] = percentile_score(
        df["debt_equity"],
        higher_is_better=False,
    )

    df["score_profit_margin"] = percentile_score(
        df["profit_margin"],
        higher_is_better=True,
    )

    df["score_operating_margin"] = percentile_score(
        df["operating_margin"],
        higher_is_better=True,
    )

    df["score_gross_margin"] = percentile_score(
        df["gross_margin"],
        higher_is_better=True,
    )

    df["score_current_ratio"] = percentile_score(
        df["current_ratio"],
        higher_is_better=True,
    )

    df["score_quick_ratio"] = percentile_score(
        df["quick_ratio"],
        higher_is_better=True,
    )

    df["score_free_cash_flow"] = percentile_score(
        df["free_cash_flow"],
        higher_is_better=True,
    )

    df["quality_score"] = (
        weighted_score(
            df,
            QUALITY_WEIGHTS,
        )
        .round(2)
    )

    return df


# ============================================================
# GROWTH SCORE
# ============================================================

def calculate_growth_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df["score_revenue_growth"] = percentile_score(
        df["revenue_growth"],
        higher_is_better=True,
    )

    df["score_earnings_growth"] = percentile_score(
        df["earnings_growth"],
        higher_is_better=True,
    )

    df[
        "score_quarterly_revenue_growth"
    ] = percentile_score(
        df["quarterly_revenue_growth"],
        higher_is_better=True,
    )

    df["growth_score"] = (
        weighted_score(
            df,
            GROWTH_WEIGHTS,
        )
        .round(2)
    )

    return df


# ============================================================
# VALUATION SCORE
# ============================================================

def calculate_valuation_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    pe = pd.to_numeric(
        df["pe"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_pe"] = percentile_score(
        pe,
        higher_is_better=False,
    )

    forward_pe = pd.to_numeric(
        df["forward_pe"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_forward_pe"] = percentile_score(
        forward_pe,
        higher_is_better=False,
    )

    pb = pd.to_numeric(
        df["price_book"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_pb"] = percentile_score(
        pb,
        higher_is_better=False,
    )

    peg = pd.to_numeric(
        df["peg"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_peg"] = percentile_score(
        peg,
        higher_is_better=False,
    )

    price_sales = pd.to_numeric(
        df["price_sales"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_price_sales"] = percentile_score(
        price_sales,
        higher_is_better=False,
    )

    ev_ebitda = pd.to_numeric(
        df["ev_ebitda"],
        errors="coerce",
    ).where(
        lambda x: x > 0
    )

    df["score_ev_ebitda"] = percentile_score(
        ev_ebitda,
        higher_is_better=False,
    )

    df["valuation_score"] = (
        weighted_score(
            df,
            VALUATION_WEIGHTS,
        )
        .round(2)
    )

    return df


# ============================================================
# FUNDAMENTAL SCORE
# ============================================================

def calculate_fundamental_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df["fundamental_score"] = (
        weighted_score(
            df,
            FUNDAMENTAL_WEIGHTS,
        )
        .round(2)
    )

    group_score_columns = [

        "quality_score",

        "growth_score",

        "valuation_score",
    ]

    df[
        "fundamental_groups_available"
    ] = (
        df[group_score_columns]
        .notna()
        .sum(axis=1)
    )

    df[
        "fundamental_group_completeness"
    ] = (
        df[
            "fundamental_groups_available"
        ]
        / len(group_score_columns)
        * 100
    ).round(2)

    if (
        "fundamental_ranking_eligible"
        not in df.columns
    ):

        df[
            "fundamental_ranking_eligible"
        ] = (
            df[
                "fundamental_confidence"
            ]
            .isin(
                [
                    "High",
                    "Medium",
                ]
            )
        )

    return df


# ============================================================
# COMBINED RESEARCH SCORE
# ============================================================

def calculate_combined_research_score(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df[
        "combined_research_score"
    ] = np.nan

    eligible = (

        df[
            "fundamental_ranking_eligible"
        ]

        &

        df[
            "market_research_score"
        ].notna()

        &

        df[
            "fundamental_score"
        ].notna()
    )

    eligible_df = df.loc[
        eligible
    ].copy()

    if not eligible_df.empty:

        df.loc[
            eligible,
            "combined_research_score",
        ] = (
            weighted_score(
                eligible_df,
                COMBINED_RESEARCH_WEIGHTS,
            )
            .round(2)
        )

    df = df.sort_values(
        by=[
            "combined_research_score",
            "market_research_score",
            "fundamental_score",
        ],
        ascending=[
            False,
            False,
            False,
        ],
        na_position="last",
    )

    df["final_rank"] = np.nan

    ranked_mask = (
        df[
            "combined_research_score"
        ].notna()
    )

    df.loc[
        ranked_mask,
        "final_rank",
    ] = np.arange(
        1,
        ranked_mask.sum() + 1,
    )

    return df


# ============================================================
# DATA QUALITY FILTER
# ============================================================

def apply_data_quality_filter(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    before = len(df)

    required_columns = [

        "symbol",

        "price",

        "data_points",

        "ma50",

        "ma200",
    ]

    df = df.dropna(
        subset=required_columns
    )

    df = df[
        (
            df["price"] >= MIN_PRICE
        )

        &

        (
            df["ma50"] > 0
        )

        &

        (
            df["ma200"] > 0
        )
    ]

    after = len(df)

    print()

    print(
        f"Stocks before data-quality filter : "
        f"{before:,}"
    )

    print(
        f"Stocks after data-quality filter  : "
        f"{after:,}"
    )

    print(
        f"Stocks removed                    : "
        f"{before - after:,}"
    )

    return df


# ============================================================
# UNIVERSE SUMMARY
# ============================================================

def display_universe_summary(
    df: pd.DataFrame,
) -> None:

    print_header(
        "UNIVERSE SUMMARY"
    )

    strong_uptrend = (
        df["strong_uptrend"].sum()
        if "strong_uptrend" in df.columns
        else 0
    )

    summary = pd.DataFrame(
        {
            "Metric": [

                "Stocks analysed",

                "Stocks above ₹100",

                "Stocks above 50 DMA",

                "Stocks above 200 DMA",

                "50 DMA above 200 DMA",

                "Stocks in strong uptrend",

                "Stocks with fundamental score",

                "High fundamental confidence",

                "Medium fundamental confidence",

                "Low fundamental confidence",

                "Eligible fundamental ranking stocks",

                "Eligible combined ranking stocks",
            ],

            "Value": [

                len(df),

                int(
                    (
                        df["price"] >= MIN_PRICE
                    ).sum()
                ),

                int(
                    df["above_ma50"].sum()
                ),

                int(
                    df["above_ma200"].sum()
                ),

                int(
                    df["ma50_above_ma200"].sum()
                ),

                int(strong_uptrend),

                int(
                    df[
                        "fundamental_score"
                    ]
                    .notna()
                    .sum()
                ),

                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "High"
                    ).sum()
                ),

                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "Medium"
                    ).sum()
                ),

                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "Low"
                    ).sum()
                ),

                int(
                    df[
                        "fundamental_ranking_eligible"
                    ].sum()
                ),

                int(
                    df[
                        "combined_research_score"
                    ]
                    .notna()
                    .sum()
                ),
            ],
        }
    )

    print(
        summary.to_string(
            index=False
        )
    )


# ============================================================
# MARKET RANKINGS
# ============================================================

def display_market_rankings(
    df: pd.DataFrame,
) -> None:

    print_header(
        f"TOP {TOP_STOCKS_TO_DISPLAY} MARKET RESEARCH RANKINGS"
    )

    market_df = (
        df
        .sort_values(
            "market_research_score",
            ascending=False,
        )
        .head(
            TOP_STOCKS_TO_DISPLAY
        )
        .copy()
    )

    if market_df.empty:

        print(
            "No market rankings available."
        )

        return

    market_df.insert(
        0,
        "market_rank",
        np.arange(
            1,
            len(market_df) + 1,
        ),
    )

    columns = [

        "market_rank",

        "symbol",

        "price",

        "return_3m",

        "return_6m",

        "return_12m",

        "price_vs_ma200",

        "distance_from_52w_high",

        "volatility",

        "max_drawdown",

        "momentum_score",

        "trend_score",

        "risk_score",

        "market_research_score",
    ]

    available = [
        column
        for column in columns
        if column in market_df.columns
    ]

    table = market_df[
        available
    ].copy()

    print(
        format_display_table(
            table
        ).to_string(
            index=False
        )
    )


# ============================================================
# FUNDAMENTAL RANKINGS
# ============================================================

def display_fundamental_rankings(
    df: pd.DataFrame,
) -> None:

    print_header(
        f"TOP {TOP_STOCKS_TO_DISPLAY} FUNDAMENTAL RANKINGS"
    )

    table_df = (
        df[
            (
                df[
                    "fundamental_ranking_eligible"
                ]
            )

            &

            (
                df[
                    "fundamental_score"
                ].notna()
            )
        ]
        .sort_values(
            [
                "fundamental_score",
                "fundamental_data_completeness",
            ],
            ascending=[
                False,
                False,
            ],
        )
        .copy()
    )

    if table_df.empty:

        print(
            "No High/Medium-confidence stocks "
            "have a valid Fundamental Score."
        )

        return

    columns = [

        "symbol",

        "price",

        "roe",
        "roa",
        "debt_equity",
        "profit_margin",
        "operating_margin",
        "gross_margin",
        "current_ratio",
        "quick_ratio",
        "free_cash_flow",

        "revenue_growth",
        "earnings_growth",
        "quarterly_revenue_growth",

        "pe",
        "forward_pe",
        "price_book",
        "peg",
        "price_sales",
        "ev_ebitda",

        "quality_score",
        "growth_score",
        "valuation_score",
        "fundamental_score",

        "quality_weighted_completeness",
        "growth_weighted_completeness",
        "valuation_weighted_completeness",
        "fundamental_data_completeness",
        "fundamental_confidence",
    ]

    available = [
        column
        for column in columns
        if column in table_df.columns
    ]

    table = (
        table_df[
            available
        ]
        .head(
            TOP_STOCKS_TO_DISPLAY
        )
        .copy()
    )

    table.insert(
        0,
        "fundamental_rank",
        np.arange(
            1,
            len(table) + 1,
        ),
    )

    print(
        format_display_table(
            table
        ).to_string(
            index=False
        )
    )


# ============================================================
# COMBINED RANKINGS
# ============================================================

def display_combined_rankings(
    df: pd.DataFrame,
) -> None:

    print_header(
        f"TOP {TOP_STOCKS_TO_DISPLAY} FUNDAMENTALALPHAFORGE RANKINGS"
    )

    table = (
        df[
            df[
                "combined_research_score"
            ].notna()
        ]
        .sort_values(
            "combined_research_score",
            ascending=False,
        )
        .head(
            TOP_STOCKS_TO_DISPLAY
        )
        .copy()
    )

    if table.empty:

        print(
            "No High/Medium-confidence stocks "
            "have a Combined Research Score."
        )

        return

    columns = [

        "final_rank",

        "symbol",

        "price",

        "market_research_score",

        "quality_score",

        "growth_score",

        "valuation_score",

        "fundamental_score",

        "quality_weighted_completeness",

        "growth_weighted_completeness",

        "valuation_weighted_completeness",

        "fundamental_data_completeness",

        "fundamental_confidence",

        "combined_research_score",
    ]

    available = [
        column
        for column in columns
        if column in table.columns
    ]

    table = table[
        available
    ]

    print(
        format_display_table(
            table
        ).to_string(
            index=False
        )
    )


# ============================================================
# DETAILED RESEARCH
# ============================================================

def display_top_detailed(
    df: pd.DataFrame,
) -> None:

    print_header(
        f"TOP {TOP_STOCKS_DETAILED} DETAILED RESEARCH"
    )

    table = (
        df[
            df[
                "combined_research_score"
            ].notna()
        ]
        .sort_values(
            "combined_research_score",
            ascending=False,
        )
        .head(
            TOP_STOCKS_DETAILED
        )
        .copy()
    )

    if table.empty:

        print(
            "No eligible stocks available "
            "for detailed research."
        )

        return

    columns = [

        "final_rank",

        "symbol",

        "price",

        "return_3m",
        "return_6m",
        "return_12m",

        "ma50",
        "ma200",

        "high_52w",
        "low_52w",
        "distance_from_52w_high",

        "volatility",
        "max_drawdown",

        "roe",
        "roa",
        "debt_equity",
        "profit_margin",
        "operating_margin",
        "gross_margin",
        "current_ratio",
        "quick_ratio",
        "free_cash_flow",

        "revenue_growth",
        "earnings_growth",
        "quarterly_revenue_growth",

        "pe",
        "forward_pe",
        "price_book",
        "peg",
        "price_sales",
        "ev_ebitda",

        "momentum_score",
        "trend_score",
        "risk_score",
        "market_research_score",

        "quality_score",
        "growth_score",
        "valuation_score",
        "fundamental_score",

        "quality_weighted_completeness",
        "growth_weighted_completeness",
        "valuation_weighted_completeness",

        "fundamental_data_completeness",

        "fundamental_confidence",

        "combined_research_score",
    ]

    available = [
        column
        for column in columns
        if column in table.columns
    ]

    table = table[
        available
    ]

    print(
        format_display_table(
            table
        ).to_string(
            index=False
        )
    )


# ============================================================
# FACTOR LEADERS
# ============================================================

def display_factor_leaders(
    df: pd.DataFrame,
) -> None:

    print_header(
        "FACTOR LEADERS"
    )

    factors = {

        "3M Return":
            "return_3m",

        "6M Return":
            "return_6m",

        "12M Return":
            "return_12m",

        "ROE":
            "roe",

        "ROA":
            "roa",

        "Profit Margin":
            "profit_margin",

        "Operating Margin":
            "operating_margin",

        "Gross Margin":
            "gross_margin",

        "Free Cash Flow":
            "free_cash_flow",

        "Revenue Growth":
            "revenue_growth",

        "Earnings Growth":
            "earnings_growth",

        "Quarterly Revenue Growth":
            "quarterly_revenue_growth",

        "Momentum Score":
            "momentum_score",

        "Trend Score":
            "trend_score",

        "Risk Score":
            "risk_score",

        "Quality Score":
            "quality_score",

        "Growth Score":
            "growth_score",

        "Valuation Score":
            "valuation_score",

        "Fundamental Score":
            "fundamental_score",

        "Combined Score":
            "combined_research_score",
    }

    rows = []

    for factor_name, column in factors.items():

        if column not in df.columns:
            continue

        factor_data = (
            df[
                [
                    "symbol",
                    column,
                ]
            ]
            .dropna()
            .sort_values(
                column,
                ascending=False,
            )
            .head(
                TOP_FACTOR_LEADERS
            )
        )

        for _, row in factor_data.iterrows():

            rows.append(
                {
                    "Factor":
                        factor_name,

                    "Symbol":
                        row["symbol"],

                    "Value":
                        round(
                            safe_float(
                                row[column]
                            ),
                            2,
                        ),
                }
            )

    if rows:

        table = pd.DataFrame(
            rows
        )

        print(
            table.to_string(
                index=False
            )
        )


# ============================================================
# RESEARCH CANDIDATES
# ============================================================

def display_research_candidates(
    df: pd.DataFrame,
) -> None:

    print_header(
        f"TOP {TOP_RESEARCH_CANDIDATES} RESEARCH CANDIDATES"
    )

    candidates = df[

        (
            df[
                "combined_research_score"
            ].notna()
        )

        &

        (
            df[
                "fundamental_ranking_eligible"
            ]
        )
    ].copy()

    candidates = candidates[
        (
            candidates[
                "market_research_score"
            ] >= 60
        )

        &

        (
            candidates[
                "fundamental_score"
            ] >= 60
        )

        &

        (
            candidates[
                "trend_score"
            ] >= 66.67
        )

        &

        (
            candidates[
                "fundamental_data_completeness"
            ] >= MEDIUM_CONFIDENCE_COMPLETENESS
        )
    ]

    candidates = candidates.sort_values(
        by=[
            "combined_research_score",
            "market_research_score",
            "fundamental_score",
        ],
        ascending=False,
    )

    candidates = candidates.head(
        TOP_RESEARCH_CANDIDATES
    )

    if candidates.empty:

        print(
            "No High/Medium-confidence stocks "
            "currently satisfy the research-candidate criteria."
        )

        return

    columns = [

        "final_rank",

        "symbol",

        "price",

        "market_research_score",

        "fundamental_score",

        "quality_score",

        "growth_score",

        "valuation_score",

        "quality_weighted_completeness",

        "growth_weighted_completeness",

        "valuation_weighted_completeness",

        "fundamental_data_completeness",

        "fundamental_confidence",

        "combined_research_score",
    ]

    available = [
        column
        for column in columns
        if column in candidates.columns
    ]

    table = candidates[
        available
    ].copy()

    print(
        format_display_table(
            table
        ).to_string(
            index=False
        )
    )


# ============================================================
# SCORE DISTRIBUTION
# ============================================================

def display_score_distribution(
    df: pd.DataFrame,
) -> None:

    print_header(
        "SCORE DISTRIBUTION"
    )

    columns = [

        "momentum_score",

        "trend_score",

        "risk_score",

        "market_research_score",

        "quality_score",

        "growth_score",

        "valuation_score",

        "fundamental_score",

        "combined_research_score",

        "quality_weighted_completeness",

        "growth_weighted_completeness",

        "valuation_weighted_completeness",

        "fundamental_data_completeness",
    ]

    available = [
        column
        for column in columns
        if column in df.columns
    ]

    if not available:
        return

    statistics = (
        df[available]
        .describe()
        .loc[
            [
                "min",
                "25%",
                "50%",
                "75%",
                "max",
            ]
        ]
        .round(2)
    )

    print(
        statistics.to_string()
    )


# ============================================================
# RESEARCH NOTES
# ============================================================

def display_research_notes() -> None:

    print_header(
        "RESEARCH NOTES"
    )

    notes = [

        "1. Current engine combines market and fundamental research.",

        "2. Market layer uses daily OHLCV data.",

        "3. Market Research Score = "
        "50% Momentum + 30% Trend + 20% Risk.",

        "4. Fundamental Score = "
        "35% Quality + 35% Growth + 30% Valuation.",

        "5. Combined Research Score = "
        "50% Market Research + 50% Fundamental.",

        "6. Quality includes profitability, margins, leverage, "
        "liquidity and free cash flow.",

        "7. Growth includes Revenue Growth, Earnings Growth "
        "and Quarterly Revenue Growth.",

        "8. EPS Growth has been removed because "
        "earningsQuarterlyGrowth is not a valid long-term EPS-growth "
        "measure.",

        "9. Valuation includes P/E, Forward P/E, P/B, PEG, "
        "Price/Sales and EV/EBITDA.",

        "10. Negative valuation multiples are treated as unavailable "
        "for valuation scoring.",

        "11. Missing fundamental values remain NaN.",

        "12. Missing fundamental values are NEVER converted to zero.",

        "13. Available factors are scored using their available weights.",

        "14. Fundamental data completeness is explicitly measured.",

        "15. Quality, Growth and Valuation completeness are measured "
        "separately.",

        "16. Weighted completeness uses the SAME raw-factor weights "
        "as the corresponding fundamental score.",

        "17. Fundamental confidence is based on weighted completeness.",

        "18. High confidence requires >=80% completeness.",

        "19. Medium confidence requires >=60% completeness.",

        "20. Low-confidence stocks are excluded from headline "
        "Fundamental and Combined rankings.",

        "21. Low-confidence stocks remain in the Excel Research Data.",

        "22. Actual factor coverage is printed for every fundamental field.",

        "23. The coverage table identifies factors exceeding 80% coverage.",

        "24. Current Yahoo Finance fundamental data is not a complete "
        "point-in-time historical dataset.",

        "25. Current fundamental research is therefore suitable for "
        "current research but not yet unbiased historical backtesting.",

        "26. Future work should add point-in-time reporting dates.",

        "27. Future work should add historical financial-statement "
        "growth measures such as 3Y Revenue CAGR and 3Y EPS CAGR.",

        "28. Future backtesting should use VectorBT after the "
        "point-in-time fundamental dataset is constructed.",

        "29. Research rankings are not BUY/SELL recommendations.",

        "30. Historical performance does not guarantee future returns.",
    ]

    for note in notes:
        print(note)


# ============================================================
# FINAL SUMMARY
# ============================================================

def display_final_summary(
    universe_count: int,
    valid_symbol_count: int,
    df: pd.DataFrame,
) -> None:

    print()

    print("=" * 80)

    print(
        "RESEARCH RUN COMPLETED"
    )

    print("=" * 80)

    print()

    print(
        f"Universe              : "
        f"{universe_count:,}"
    )

    print(
        f"Valid market symbols  : "
        f"{valid_symbol_count:,}"
    )

    print(
        f"Market-researched     : "
        f"{len(df):,}"
    )

    fundamental_count = int(
        df[
            "fundamental_score"
        ]
        .notna()
        .sum()
    )

    print(
        f"Fundamental scores    : "
        f"{fundamental_count:,}"
    )

    eligible_count = int(
        df[
            "fundamental_ranking_eligible"
        ].sum()
    )

    print(
        f"Fundamental eligible  : "
        f"{eligible_count:,}"
    )

    combined_count = int(
        df[
            "combined_research_score"
        ]
        .notna()
        .sum()
    )

    print(
        f"Combined ranking      : "
        f"{combined_count:,}"
    )

    if (
        "fundamental_data_completeness"
        in df.columns
    ):

        print(
            f"Avg fundamental data : "
            f"{df['fundamental_data_completeness'].mean():.2f}%"
        )

        print(
            f"Median fundamental   : "
            f"{df['fundamental_data_completeness'].median():.2f}%"
        )

    eligible_df = df[
        df[
            "combined_research_score"
        ].notna()
    ].copy()

    if not eligible_df.empty:

        best = (
            eligible_df
            .sort_values(
                "combined_research_score",
                ascending=False,
            )
            .iloc[0]
        )

        print()

        print(
            "TOP FUNDAMENTALALPHAFORGE RESEARCH STOCK:"
        )

        print(
            f"  {best['symbol']}"
        )

        print(
            f"  Combined Score      : "
            f"{best['combined_research_score']:.2f}"
        )

        print(
            f"  Market Score        : "
            f"{best['market_research_score']:.2f}"
        )

        print(
            f"  Fundamental Score   : "
            f"{best['fundamental_score']:.2f}"
        )

        print(
            f"  Quality Score       : "
            f"{best['quality_score']:.2f}"
        )

        print(
            f"  Growth Score        : "
            f"{best['growth_score']:.2f}"
        )

        print(
            f"  Valuation Score     : "
            f"{best['valuation_score']:.2f}"
        )

        print(
            f"  Fundamental Data    : "
            f"{best['fundamental_data_completeness']:.2f}%"
        )

        print(
            f"  Data Confidence     : "
            f"{best['fundamental_confidence']}"
        )

    print()

    print(
        "Fundamental model:"
    )

    print(
        "  Quality   = profitability + margins + "
        "leverage + liquidity + cash flow"
    )

    print(
        "  Growth    = revenue + earnings + "
        "quarterly revenue growth"
    )

    print(
        "  Valuation = earnings + book + sales + "
        "enterprise-value based measures"
    )

    print()

    print(
        "Current pipeline:"
    )

    print(
        "Nifty 500"
        " → Market Data"
        " → Momentum"
        " → Trend"
        " → Risk"
        " → Quality"
        " → Growth"
        " → Valuation"
        " → Completeness"
        " → Confidence"
        " → Combined Ranking"
        " → Excel Dashboard"
    )

    print()

    print(
        "Next major research stage:"
    )

    print(
        "Historical Financial Statements"
        " → TTM / 3Y / 5Y Growth"
        " → Point-in-Time Dates"
        " → VectorBT"
        " → Walk-Forward Testing"
        " → Out-of-Sample Validation"
    )

    print()


# ============================================================
# EXCEL HELPERS
# ============================================================

def excel_safe_value(value):

    if isinstance(value, (np.integer,)):

        return int(value)

    if isinstance(value, (np.floating,)):

        if np.isnan(value) or np.isinf(value):

            return None

        return float(value)

    if pd.isna(value):

        return None

    if isinstance(value, (np.bool_,)):

        return bool(value)

    return value


def write_dataframe_to_sheet(
    workbook,
    dataframe: pd.DataFrame,
    sheet_name: str,
) -> None:

    if sheet_name in workbook.sheetnames:

        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name
    )

    for column_index, column in enumerate(
        dataframe.columns,
        start=1,
    ):

        cell = ws.cell(
            row=1,
            column=column_index,
            value=str(column),
        )

        cell.font = Font(
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_index, row in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        ),
        start=2,
    ):

        for column_index, value in enumerate(
            row,
            start=1,
        ):

            ws.cell(
                row=row_index,
                column=column_index,
                value=excel_safe_value(value),
            )

    ws.freeze_panes = "A2"

    if ws.max_row > 1 and ws.max_column > 0:

        ws.auto_filter.ref = ws.dimensions

    # --------------------------------------------------------
    # Header formatting
    # --------------------------------------------------------

    for cell in ws[1]:

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78",
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

    # --------------------------------------------------------
    # Number formatting
    # --------------------------------------------------------

    for row in ws.iter_rows(
        min_row=2,
    ):

        for cell in row:

            if isinstance(
                cell.value,
                float,
            ):

                cell.number_format = "0.00"

    # --------------------------------------------------------
    # Widths
    # --------------------------------------------------------

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells[:500]:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            28,
        )


def style_dashboard_title(
    ws,
    cell_range: str,
    title: str,
) -> None:

    ws.merge_cells(
        cell_range
    )

    cell = ws[
        cell_range.split(":")[0]
    ]

    cell.value = title

    cell.font = Font(
        bold=True,
        size=18,
        color="FFFFFF",
    )

    cell.fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )


def style_section(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    title: str,
) -> None:

    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )

    cell = ws.cell(
        row=row,
        column=start_col,
    )

    cell.value = title

    cell.font = Font(
        bold=True,
        size=12,
        color="FFFFFF",
    )

    cell.fill = PatternFill(
        "solid",
        fgColor="4472C4",
    )

    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )


def create_kpi(
    ws,
    row: int,
    start_col: int,
    title: str,
    value,
) -> None:

    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=start_col + 2,
    )

    ws.merge_cells(
        start_row=row + 1,
        start_column=start_col,
        end_row=row + 2,
        end_column=start_col + 2,
    )

    title_cell = ws.cell(
        row=row,
        column=start_col,
    )

    value_cell = ws.cell(
        row=row + 1,
        column=start_col,
    )

    title_cell.value = title

    value_cell.value = value

    title_cell.font = Font(
        bold=True,
        color="FFFFFF",
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor="5B9BD5",
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    value_cell.font = Font(
        bold=True,
        size=18,
        color="1F1F1F",
    )

    value_cell.fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )


def write_table(
    ws,
    dataframe: pd.DataFrame,
    start_row: int,
    start_col: int,
    title: Optional[str] = None,
) -> int:

    row = start_row

    if dataframe is None:

        return row

    if dataframe.empty:

        ws.cell(
            row=row,
            column=start_col,
            value=(
                title
                if title
                else "No data available"
            ),
        )

        ws.cell(
            row=row,
            column=start_col,
        ).font = Font(
            bold=True,
            size=11,
        )

        return row + 1

    if title:

        ws.cell(
            row=row,
            column=start_col,
            value=title,
        )

        ws.cell(
            row=row,
            column=start_col,
        ).font = Font(
            bold=True,
            size=11,
        )

        row += 1

    for col_index, column in enumerate(
        dataframe.columns,
        start=start_col,
    ):

        cell = ws.cell(
            row=row,
            column=col_index,
            value=str(column),
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="5B9BD5",
        )

        cell.alignment = Alignment(
            horizontal="center",
        )

    header_row = row

    row += 1

    for values in dataframe.itertuples(
        index=False,
        name=None,
    ):

        for col_index, value in enumerate(
            values,
            start=start_col,
        ):

            cell = ws.cell(
                row=row,
                column=col_index,
                value=excel_safe_value(value),
            )

            if isinstance(
                cell.value,
                float,
            ):

                cell.number_format = "0.00"

        row += 1

    # --------------------------------------------------------
    # Borders
    # --------------------------------------------------------

    thin = Side(
        style="thin",
        color="D9E1F2",
    )

    for r in ws.iter_rows(
        min_row=header_row,
        max_row=row - 1,
        min_col=start_col,
        max_col=start_col + len(dataframe.columns) - 1,
    ):

        for cell in r:

            cell.border = Border(
                bottom=thin,
            )

    return row


def add_bar_chart(
    ws,
    data_start_row: int,
    data_end_row: int,
    category_col: int,
    value_col: int,
    title: str,
    anchor: str,
    horizontal: bool = False,
) -> None:

    if data_end_row < data_start_row:

        return

    chart = BarChart()

    if horizontal:

        chart.type = "bar"

    else:

        chart.type = "col"

    chart.style = 10

    chart.title = title

    chart.y_axis.title = "Value"

    chart.x_axis.title = "Category"

    data = Reference(
        ws,
        min_col=value_col,
        min_row=data_start_row - 1,
        max_row=data_end_row,
    )

    categories = Reference(
        ws,
        min_col=category_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(
        categories
    )

    chart.height = 8

    chart.width = 15

    chart.legend = None

    ws.add_chart(
        chart,
        anchor,
    )


def add_pie_chart(
    ws,
    data_start_row: int,
    data_end_row: int,
    category_col: int,
    value_col: int,
    title: str,
    anchor: str,
) -> None:

    if data_end_row < data_start_row:

        return

    chart = PieChart()

    chart.title = title

    data = Reference(
        ws,
        min_col=value_col,
        min_row=data_start_row - 1,
        max_row=data_end_row,
    )

    labels = Reference(
        ws,
        min_col=category_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(
        labels
    )

    chart.height = 8

    chart.width = 12

    chart.dataLabels = DataLabelList()

    chart.dataLabels.showPercent = True

    ws.add_chart(
        chart,
        anchor,
    )


def add_scatter_chart(
    ws,
    x_col: int,
    y_col: int,
    start_row: int,
    end_row: int,
    title: str,
    anchor: str,
) -> None:

    if end_row < start_row:

        return

    chart = ScatterChart()

    chart.title = title

    chart.x_axis.title = str(
        ws.cell(
            row=1,
            column=x_col,
        ).value
    )

    chart.y_axis.title = str(
        ws.cell(
            row=1,
            column=y_col,
        ).value
    )

    chart.height = 9

    chart.width = 14

    xvalues = Reference(
        ws,
        min_col=x_col,
        min_row=start_row,
        max_row=end_row,
    )

    yvalues = Reference(
        ws,
        min_col=y_col,
        min_row=start_row,
        max_row=end_row,
    )

    series = Series(
        yvalues,
        xvalues,
        title="Stocks",
    )

    chart.series.append(
        series
    )

    ws.add_chart(
        chart,
        anchor,
    )


# ============================================================
# BUILD FACTOR COVERAGE SHEET
# ============================================================

def build_factor_coverage_sheet(
    workbook,
    df: pd.DataFrame,
) -> None:

    if COVERAGE_SHEET in workbook.sheetnames:

        del workbook[COVERAGE_SHEET]

    ws = workbook.create_sheet(
        COVERAGE_SHEET
    )

    style_dashboard_title(
        ws,
        "A1:G1",
        "Fundamental Factor Coverage",
    )

    rows = []

    total = len(df)

    for factor in ALL_FUNDAMENTAL_FACTORS:

        if factor not in df.columns:
            continue

        available = int(
            df[factor].notna().sum()
        )

        missing = total - available

        coverage = (
            available / total * 100
            if total
            else 0
        )

        if coverage >= 80:

            status = "Strong"

        elif coverage >= 60:

            status = "Moderate"

        else:

            status = "Sparse"

        if factor in FUNDAMENTAL_FACTOR_GROUPS["quality"]:

            group = "Quality"

        elif factor in FUNDAMENTAL_FACTOR_GROUPS["growth"]:

            group = "Growth"

        else:

            group = "Valuation"

        rows.append(
            {
                "Group": group,
                "Factor": factor,
                "Available": available,
                "Missing": missing,
                "Coverage %": round(
                    coverage,
                    2,
                ),
                "Status": status,
                "Weight": (
                    QUALITY_FACTOR_WEIGHTS.get(
                        factor,
                        GROWTH_FACTOR_WEIGHTS.get(
                            factor,
                            VALUATION_FACTOR_WEIGHTS.get(
                                factor,
                                np.nan,
                            ),
                        ),
                    )
                ),
            }
        )

    coverage_df = pd.DataFrame(
        rows
    )

    write_table(
        ws,
        coverage_df,
        start_row=3,
        start_col=1,
    )

    ws.freeze_panes = "A4"

    for row in range(
        4,
        4 + len(coverage_df),
    ):

        ws.cell(
            row=row,
            column=5,
        ).number_format = "0.00"

        ws.cell(
            row=row,
            column=7,
        ).number_format = "0.00"

    if len(coverage_df) > 0:

        ws.conditional_formatting.add(
            f"E4:E{3 + len(coverage_df)}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

        # ----------------------------------------------------
        # IMPORTANT FIX
        #
        # Explicitly specify the DataBar color.
        #
        # Without color=..., some openpyxl versions create
        # DataBar.color as None and fail with:
        #
        # <class 'openpyxl.formatting.rule.DataBar'>.color
        # should be <class 'openpyxl.styles.colors.Color'>
        # but value is <class 'NoneType'>
        # ----------------------------------------------------

        ws.conditional_formatting.add(
            f"E4:E{3 + len(coverage_df)}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=100,
                color="5B9BD5",
                showValue=True,
            ),
        )

    for column in range(1, 8):

        ws.column_dimensions[
            get_column_letter(column)
        ].width = 20


# ============================================================
# BUILD SCORE STATISTICS SHEET
# ============================================================

def build_score_statistics_sheet(
    workbook,
    df: pd.DataFrame,
) -> None:

    if STATISTICS_SHEET in workbook.sheetnames:

        del workbook[STATISTICS_SHEET]

    ws = workbook.create_sheet(
        STATISTICS_SHEET
    )

    style_dashboard_title(
        ws,
        "A1:H1",
        "Score & Research Statistics",
    )

    score_columns = [

        "momentum_score",
        "trend_score",
        "risk_score",
        "market_research_score",
        "quality_score",
        "growth_score",
        "valuation_score",
        "fundamental_score",
        "combined_research_score",
        "fundamental_data_completeness",
    ]

    available = [
        c
        for c in score_columns
        if c in df.columns
    ]

    if not available:

        ws["A3"] = "No score columns available."

        return

    statistics = (
        df[available]
        .describe()
        .T
        .reset_index()
        .rename(
            columns={
                "index": "Metric",
            }
        )
    )

    write_table(
        ws,
        statistics.round(2),
        start_row=3,
        start_col=1,
    )

    # --------------------------------------------------------
    # Correlation matrix
    # --------------------------------------------------------

    corr_start = 7

    style_section(
        ws,
        corr_start,
        1,
        min(
            len(available) + 1,
            12,
        ),
        "Score Correlation Matrix",
    )

    corr = (
        df[available]
        .corr()
        .round(2)
    )

    for col_idx, column in enumerate(
        corr.columns,
        start=1,
    ):

        ws.cell(
            row=corr_start + 1,
            column=col_idx + 1,
            value=column,
        )

        ws.cell(
            row=corr_start + 1,
            column=col_idx + 1,
        ).font = Font(
            bold=True,
        )

    for r, row_name in enumerate(
        corr.index,
        start=corr_start + 2,
    ):

        ws.cell(
            row=r,
            column=1,
            value=row_name,
        )

        for c, col_name in enumerate(
            corr.columns,
            start=2,
        ):

            ws.cell(
                row=r,
                column=c,
                value=excel_safe_value(
                    corr.loc[
                        row_name,
                        col_name,
                    ]
                ),
            )

    end_row = (
        corr_start
        + 1
        + len(corr)
    )

    end_col = (
        1
        + len(corr.columns)
    )

    if len(corr) > 0:

        ws.conditional_formatting.add(
            f"B{corr_start + 2}:{get_column_letter(end_col)}{end_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    for column in range(
        1,
        max(
            10,
            end_col + 1,
        ),
    ):

        ws.column_dimensions[
            get_column_letter(column)
        ].width = 24


# ============================================================
# BUILD EXCEL DASHBOARD
# ============================================================

def build_excel_dashboard(
    df: pd.DataFrame,
    universe_count: int,
    valid_symbol_count: int,
    output_file: Path,
) -> None:

    print_header(
        "BUILDING EXCEL DASHBOARD"
    )

    RESULTS_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # --------------------------------------------------------
    # Create workbook
    # --------------------------------------------------------

    from openpyxl import Workbook

    workbook = Workbook()

    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    # --------------------------------------------------------
    # Research Data
    # --------------------------------------------------------

    write_dataframe_to_sheet(
        workbook,
        df,
        RESEARCH_DATA_SHEET,
    )

    # --------------------------------------------------------
    # Coverage
    # --------------------------------------------------------

    build_factor_coverage_sheet(
        workbook,
        df,
    )

    # --------------------------------------------------------
    # Statistics
    # --------------------------------------------------------

    build_score_statistics_sheet(
        workbook,
        df,
    )

    # --------------------------------------------------------
    # Dashboard
    # --------------------------------------------------------

    ws = workbook.create_sheet(
        DASHBOARD_SHEET,
        0,
    )

    ws.sheet_view.showGridLines = False

    style_dashboard_title(
        ws,
        "A1:N2",
        "FUNDAMENTALALPHAFORGE — EQUITY RESEARCH DASHBOARD",
    )

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # ========================================================
    # RUN INFORMATION
    # ========================================================

    style_section(
        ws,
        4,
        1,
        14,
        "RUN INFORMATION",
    )

    run_start = (
        PROGRAM_START_TIME.strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        if PROGRAM_START_TIME
        else "N/A"
    )

    run_end = (
        PROGRAM_END_TIME.strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        if PROGRAM_END_TIME
        else "N/A"
    )

    elapsed = (
        PROGRAM_ELAPSED_SECONDS
        if PROGRAM_ELAPSED_SECONDS is not None
        else np.nan
    )

    ws["A5"] = "Start Timestamp"
    ws["B5"] = run_start

    ws["D5"] = "End Timestamp"
    ws["E5"] = run_end

    ws["G5"] = "Runtime Seconds"

    ws["H5"] = (
        round(
            elapsed,
            2,
        )
        if pd.notna(elapsed)
        else "N/A"
    )

    ws["J5"] = "Runtime"

    ws["K5"] = (
        time.strftime(
            "%H:%M:%S",
            time.gmtime(
                elapsed
            ),
        )
        if pd.notna(elapsed)
        else "N/A"
    )

    for cell in [
        "A5",
        "D5",
        "G5",
        "J5",
    ]:

        ws[cell].font = Font(
            bold=True,
        )

    # ========================================================
    # KPI CARDS
    # ========================================================

    style_section(
        ws,
        7,
        1,
        14,
        "KEY RESEARCH INDICATORS",
    )

    stocks_analysed = len(df)

    combined_count = int(
        df[
            "combined_research_score"
        ].notna().sum()
    )

    high_confidence = int(
        (
            df[
                "fundamental_confidence"
            ] == "High"
        ).sum()
    )

    medium_confidence = int(
        (
            df[
                "fundamental_confidence"
            ] == "Medium"
        ).sum()
    )

    strong_uptrend = int(
        df["strong_uptrend"].fillna(False).sum()
    )

    avg_completeness = (
        df[
            "fundamental_data_completeness"
        ].mean()
    )

    median_completeness = (
        df[
            "fundamental_data_completeness"
        ].median()
    )

    create_kpi(
        ws,
        8,
        1,
        "Universe",
        universe_count,
    )

    create_kpi(
        ws,
        8,
        4,
        "Market Researched",
        stocks_analysed,
    )

    create_kpi(
        ws,
        8,
        7,
        "Combined Eligible",
        combined_count,
    )

    create_kpi(
        ws,
        8,
        10,
        "Strong Uptrend",
        strong_uptrend,
    )

    create_kpi(
        ws,
        12,
        1,
        "High Confidence",
        high_confidence,
    )

    create_kpi(
        ws,
        12,
        4,
        "Medium Confidence",
        medium_confidence,
    )

    create_kpi(
        ws,
        12,
        7,
        "Avg Completeness",
        (
            f"{avg_completeness:.2f}%"
            if pd.notna(avg_completeness)
            else "N/A"
        ),
    )

    create_kpi(
        ws,
        12,
        10,
        "Median Completeness",
        (
            f"{median_completeness:.2f}%"
            if pd.notna(median_completeness)
            else "N/A"
        ),
    )

    # ========================================================
    # TOP RESEARCH STOCKS
    # ========================================================

    style_section(
        ws,
        17,
        1,
        7,
        "TOP 10 COMBINED RESEARCH STOCKS",
    )

    top_combined_columns = [
        "final_rank",
        "symbol",
        "price",
        "market_research_score",
        "fundamental_score",
        "fundamental_data_completeness",
        "combined_research_score",
    ]

    top_combined = (
        df[
            df[
                "combined_research_score"
            ].notna()
        ]
        .sort_values(
            "combined_research_score",
            ascending=False,
        )
        .head(10)
        [
            [
                c
                for c in top_combined_columns
                if c in df.columns
            ]
        ]
        .copy()
    )

    write_table(
        ws,
        top_combined.round(2),
        18,
        1,
    )

    # ========================================================
    # TOP MARKET STOCKS
    # ========================================================

    style_section(
        ws,
        17,
        9,
        14,
        "TOP 10 MARKET STOCKS",
    )

    top_market_columns = [
        "symbol",
        "price",
        "momentum_score",
        "trend_score",
        "risk_score",
        "market_research_score",
    ]

    top_market = (
        df
        .sort_values(
            "market_research_score",
            ascending=False,
        )
        .head(10)
        [
            [
                c
                for c in top_market_columns
                if c in df.columns
            ]
        ]
        .copy()
    )

    write_table(
        ws,
        top_market.round(2),
        18,
        9,
    )

    # ========================================================
    # CONFIDENCE BREAKDOWN
    # ========================================================

    confidence_row = 34

    style_section(
        ws,
        confidence_row,
        1,
        6,
        "FUNDAMENTAL CONFIDENCE BREAKDOWN",
    )

    confidence_df = pd.DataFrame(
        {
            "Confidence": [
                "High",
                "Medium",
                "Low",
                "No Data",
            ],
            "Stocks": [
                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "High"
                    ).sum()
                ),
                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "Medium"
                    ).sum()
                ),
                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "Low"
                    ).sum()
                ),
                int(
                    (
                        df[
                            "fundamental_confidence"
                        ] == "No Data"
                    ).sum()
                ),
            ],
        }
    )

    write_table(
        ws,
        confidence_df,
        confidence_row + 1,
        1,
    )

    add_pie_chart(
        ws,
        confidence_row + 2,
        confidence_row + 5,
        1,
        2,
        "Fundamental Confidence",
        "D34",
    )

    # ========================================================
    # MARKET SCORE COMPONENTS
    # ========================================================

    component_row = 42

    style_section(
        ws,
        component_row,
        1,
        7,
        "AVERAGE MARKET SCORE COMPONENTS",
    )

    market_component_df = pd.DataFrame(
        {
            "Component": [
                "Momentum",
                "Trend",
                "Risk",
                "Market Research",
            ],
            "Average Score": [
                round(
                    df["momentum_score"].mean(),
                    2,
                ),
                round(
                    df["trend_score"].mean(),
                    2,
                ),
                round(
                    df["risk_score"].mean(),
                    2,
                ),
                round(
                    df["market_research_score"].mean(),
                    2,
                ),
            ],
        }
    )

    write_table(
        ws,
        market_component_df,
        component_row + 1,
        1,
    )

    add_bar_chart(
        ws,
        component_row + 2,
        component_row + 5,
        1,
        2,
        "Market Score Components",
        "D42",
    )

    # ========================================================
    # FUNDAMENTAL COMPONENTS
    # ========================================================

    fundamental_row = 42

    style_section(
        ws,
        fundamental_row,
        9,
        14,
        "AVERAGE FUNDAMENTAL COMPONENTS",
    )

    fundamental_component_df = pd.DataFrame(
        {
            "Component": [
                "Quality",
                "Growth",
                "Valuation",
                "Fundamental",
            ],
            "Average Score": [
                round(
                    df["quality_score"].mean(),
                    2,
                ),
                round(
                    df["growth_score"].mean(),
                    2,
                ),
                round(
                    df["valuation_score"].mean(),
                    2,
                ),
                round(
                    df["fundamental_score"].mean(),
                    2,
                ),
            ],
        }
    )

    write_table(
        ws,
        fundamental_component_df,
        fundamental_row + 1,
        9,
    )

    add_bar_chart(
        ws,
        fundamental_row + 2,
        fundamental_row + 5,
        9,
        10,
        "Fundamental Score Components",
        "L42",
    )

    # ========================================================
    # FACTOR COVERAGE
    # ========================================================

    coverage_row = 50

    style_section(
        ws,
        coverage_row,
        1,
        7,
        "FUNDAMENTAL FACTOR COVERAGE",
    )

    coverage_rows = []

    for factor in ALL_FUNDAMENTAL_FACTORS:

        if factor not in df.columns:
            continue

        coverage = (
            df[factor].notna().mean()
            * 100
        )

        coverage_rows.append(
            {
                "Factor": factor,
                "Coverage %": round(
                    coverage,
                    2,
                ),
            }
        )

    coverage_dashboard_df = (
        pd.DataFrame(
            coverage_rows
        )
        .sort_values(
            "Coverage %",
            ascending=False,
        )
        .head(20)
    )

    write_table(
        ws,
        coverage_dashboard_df,
        coverage_row + 1,
        1,
    )

    if not coverage_dashboard_df.empty:

        add_bar_chart(
            ws,
            coverage_row + 2,
            coverage_row + 1 + len(
                coverage_dashboard_df
            ),
            1,
            2,
            "Fundamental Factor Coverage",
            "D50",
            horizontal=True,
        )

    # ========================================================
    # RISK / RETURN TABLE
    # ========================================================

    risk_row = 50

    style_section(
        ws,
        risk_row,
        9,
        14,
        "RISK / RETURN STATISTICS",
    )

    risk_return_df = pd.DataFrame(
        {
            "Metric": [
                "Average 3M Return",
                "Median 3M Return",
                "Average 6M Return",
                "Median 6M Return",
                "Average 12M Return",
                "Median 12M Return",
                "Average Volatility",
                "Median Volatility",
                "Average Max Drawdown",
                "Median Max Drawdown",
            ],
            "Value": [
                df["return_3m"].mean(),
                df["return_3m"].median(),
                df["return_6m"].mean(),
                df["return_6m"].median(),
                df["return_12m"].mean(),
                df["return_12m"].median(),
                df["volatility"].mean(),
                df["volatility"].median(),
                df["max_drawdown"].mean(),
                df["max_drawdown"].median(),
            ],
        }
    )

    write_table(
        ws,
        risk_return_df.round(2),
        risk_row + 1,
        9,
    )

    # ========================================================
    # TOP QUALITY
    # ========================================================

    quality_row = 64

    style_section(
        ws,
        quality_row,
        1,
        7,
        "TOP 10 QUALITY STOCKS",
    )

    quality_columns = [
        "symbol",
        "quality_score",
        "roe",
        "roa",
        "profit_margin",
        "operating_margin",
        "debt_equity",
    ]

    top_quality = (
        df[
            df["quality_score"].notna()
        ]
        .sort_values(
            "quality_score",
            ascending=False,
        )
        .head(10)
        [
            [
                c
                for c in quality_columns
                if c in df.columns
            ]
        ]
        .copy()
    )

    write_table(
        ws,
        top_quality.round(2),
        quality_row + 1,
        1,
    )

    # ========================================================
    # TOP GROWTH
    # ========================================================

    growth_row = 64

    style_section(
        ws,
        growth_row,
        9,
        14,
        "TOP 10 GROWTH STOCKS",
    )

    growth_columns = [
        "symbol",
        "growth_score",
        "revenue_growth",
        "earnings_growth",
        "quarterly_revenue_growth",
    ]

    top_growth = (
        df[
            df["growth_score"].notna()
        ]
        .sort_values(
            "growth_score",
            ascending=False,
        )
        .head(10)
        [
            [
                c
                for c in growth_columns
                if c in df.columns
            ]
        ]
        .copy()
    )

    write_table(
        ws,
        top_growth.round(2),
        growth_row + 1,
        9,
    )

    # ========================================================
    # TOP VALUATION
    # ========================================================

    valuation_row = 78

    style_section(
        ws,
        valuation_row,
        1,
        7,
        "TOP 10 VALUATION SCORES",
    )

    valuation_columns = [
        "symbol",
        "valuation_score",
        "pe",
        "forward_pe",
        "price_book",
        "peg",
        "ev_ebitda",
    ]

    top_valuation = (
        df[
            df["valuation_score"].notna()
        ]
        .sort_values(
            "valuation_score",
            ascending=False,
        )
        .head(10)
        [
            [
                c
                for c in valuation_columns
                if c in df.columns
            ]
        ]
        .copy()
    )

    write_table(
        ws,
        top_valuation.round(2),
        valuation_row + 1,
        1,
    )

    # ========================================================
    # RESEARCH CANDIDATES
    # ========================================================

    candidate_row = 78

    style_section(
        ws,
        candidate_row,
        9,
        14,
        "RESEARCH CANDIDATE SUMMARY",
    )

    candidates = df[

        (
            df[
                "combined_research_score"
            ].notna()
        )

        &

        (
            df[
                "fundamental_ranking_eligible"
            ].fillna(False)
        )

        &

        (
            df[
                "market_research_score"
            ] >= 60
        )

        &

        (
            df[
                "fundamental_score"
            ] >= 60
        )

        &

        (
            df[
                "trend_score"
            ] >= 66.67
        )

        &

        (
            df[
                "fundamental_data_completeness"
            ] >= MEDIUM_CONFIDENCE_COMPLETENESS
        )
    ]

    candidate_summary = pd.DataFrame(
        {
            "Metric": [
                "Research candidates",
                "Market score >= 60",
                "Fundamental score >= 60",
                "Trend score >= 66.67",
                "Completeness >= 60%",
            ],
            "Count": [
                len(candidates),
                int(
                    (
                        df[
                            "market_research_score"
                        ] >= 60
                    ).sum()
                ),
                int(
                    (
                        df[
                            "fundamental_score"
                        ] >= 60
                    ).sum()
                ),
                int(
                    (
                        df[
                            "trend_score"
                        ] >= 66.67
                    ).sum()
                ),
                int(
                    (
                        df[
                            "fundamental_data_completeness"
                        ] >= 60
                    ).sum()
                ),
            ],
        }
    )

    write_table(
        ws,
        candidate_summary,
        candidate_row + 1,
        9,
    )

    # ========================================================
    # CONDITIONAL FORMATTING — RESEARCH DATA
    # ========================================================

    research_ws = workbook[
        RESEARCH_DATA_SHEET
    ]

    header_map = {}

    for cell in research_ws[1]:

        header_map[
            cell.value
        ] = cell.column

    score_columns = [

        "momentum_score",
        "trend_score",
        "risk_score",
        "market_research_score",
        "quality_score",
        "growth_score",
        "valuation_score",
        "fundamental_score",
        "fundamental_data_completeness",
        "combined_research_score",
    ]

    for column in score_columns:

        if column not in header_map:

            continue

        col = get_column_letter(
            header_map[column]
        )

        if research_ws.max_row >= 2:

            research_ws.conditional_formatting.add(
                f"{col}2:{col}{research_ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    # ========================================================
    # FINAL DASHBOARD FORMATTING
    # ========================================================

    for column in range(
        1,
        15,
    ):

        ws.column_dimensions[
            get_column_letter(column)
        ].width = 16

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 18

    ws.freeze_panes = "A4"

    # ========================================================
    # SAVE
    # ========================================================

    workbook.save(
        output_file
    )

    print()

    print(
        f"Excel workbook created: {output_file}"
    )

    print(
        "Sheets created:"
    )

    print(
        f"  1. {DASHBOARD_SHEET}"
    )

    print(
        f"  2. {RESEARCH_DATA_SHEET}"
    )

    print(
        f"  3. {COVERAGE_SHEET}"
    )

    print(
        f"  4. {STATISTICS_SHEET}"
    )


# ============================================================
# SAVE RESULTS TO EXCEL
# ============================================================

def save_results(
    df: pd.DataFrame,
) -> Path:

    RESULTS_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    build_excel_dashboard(
        df=df,
        universe_count=len(df),
        valid_symbol_count=len(df),
        output_file=EXCEL_OUTPUT_FILE,
    )

    return EXCEL_OUTPUT_FILE


# ============================================================
# MAIN
# ============================================================

def main() -> None:

    global PROGRAM_START_TIME
    global PROGRAM_END_TIME
    global PROGRAM_ELAPSED_SECONDS

    # ========================================================
    # START TIMESTAMP
    # ========================================================

    PROGRAM_START_TIME = datetime.now()

    performance_start = time.perf_counter()

    print()

    print("=" * 80)

    print(
        "              FUNDAMENTALALPHAFORGE"
    )

    print(
        "          QUANTITATIVE EQUITY RESEARCH"
    )

    print("=" * 80)

    print()

    print(
        f"Program started : "
        f"{PROGRAM_START_TIME.strftime('%Y-%m-%d %H:%M:%S')}"
    )

    print()

    print(
        f"Project root: {PROJECT_ROOT}"
    )

    # ========================================================
    # STEP 1
    # UNIVERSE
    # ========================================================

    symbols = (
        refresh_and_load_universe()
    )

    # ========================================================
    # STEP 2
    # MARKET DATA
    # ========================================================

    print_header(
        "MARKET DATA DOWNLOAD"
    )

    print(
        f"Historical period : "
        f"{HISTORICAL_PERIOD}"
    )

    print(
        f"Symbols requested : "
        f"{len(symbols):,}"
    )

    print()

    print(
        "Downloading market data..."
    )

    suppressed_output = io.StringIO()

    with redirect_stdout(
        suppressed_output
    ):

        data, valid_symbols = (
            get_historical_market_data_for_symbols(
                symbols,
                period=HISTORICAL_PERIOD,
            )
        )

    print(
        "Market data download completed."
    )

    if data is None or data.empty:

        raise RuntimeError(
            "Yahoo Finance returned no market data."
        )

    if not valid_symbols:

        raise RuntimeError(
            "No valid symbols returned from Yahoo Finance."
        )

    print()

    print(
        f"Historical data period: "
        f"{HISTORICAL_PERIOD}"
    )

    print(
        f"Requested symbols: "
        f"{len(symbols):,}"
    )

    print(
        f"Valid symbols: "
        f"{len(valid_symbols):,}"
    )

    print(
        f"Invalid symbols: "
        f"{len(symbols) - len(valid_symbols):,}"
    )

    # ========================================================
    # STEP 3
    # MARKET METRICS
    # ========================================================

    df = build_research_dataframe(
        data,
        valid_symbols,
    )

    # ========================================================
    # STEP 4
    # DATA QUALITY
    # ========================================================

    df = apply_data_quality_filter(
        df
    )

    if df.empty:

        raise RuntimeError(
            "All stocks were removed by the data-quality filter."
        )

    # ========================================================
    # STEP 5
    # MOMENTUM
    # ========================================================

    print_header(
        "CALCULATING MOMENTUM SCORE"
    )

    df = calculate_momentum_score(
        df
    )

    print(
        "Momentum Score calculated."
    )

    # ========================================================
    # STEP 6
    # TREND
    # ========================================================

    print_header(
        "CALCULATING TREND SCORE"
    )

    df = calculate_trend_score(
        df
    )

    print(
        "Trend Score calculated."
    )

    # ========================================================
    # STEP 7
    # RISK
    # ========================================================

    print_header(
        "CALCULATING RISK SCORE"
    )

    df = calculate_risk_score(
        df
    )

    print(
        "Risk Score calculated."
    )

    # ========================================================
    # STEP 8
    # MARKET RESEARCH
    # ========================================================

    print_header(
        "CALCULATING MARKET RESEARCH SCORE"
    )

    df = calculate_market_research_score(
        df
    )

    print(
        "Market Research Score calculated."
    )

    # ========================================================
    # STEP 9
    # FUNDAMENTALS
    # ========================================================

    fundamentals_df = (
        build_fundamental_dataframe(
            symbols
        )
    )

    # ========================================================
    # STEP 10
    # MERGE
    # ========================================================

    if not fundamentals_df.empty:

        df = df.merge(
            fundamentals_df,
            on="symbol",
            how="left",
        )

    else:

        fundamental_columns = [

            "roe",
            "roa",
            "debt_equity",

            "profit_margin",
            "operating_margin",
            "gross_margin",
            "current_ratio",
            "quick_ratio",
            "free_cash_flow",

            "revenue_growth",
            "earnings_growth",
            "quarterly_revenue_growth",

            "pe",
            "forward_pe",
            "price_book",
            "peg",
            "price_sales",
            "ev_ebitda",

            "market_cap",
            "enterprise_value",
        ]

        for column in fundamental_columns:

            df[column] = np.nan

    # ========================================================
    # STEP 11
    # FUNDAMENTAL DATA QUALITY
    # ========================================================

    df = calculate_fundamental_data_quality(
        df
    )

    # ========================================================
    # STEP 12
    # AVAILABILITY
    # ========================================================

    display_fundamental_availability(
        df
    )

    # ========================================================
    # STEP 13
    # QUALITY
    # ========================================================

    print_header(
        "CALCULATING QUALITY SCORE"
    )

    df = calculate_quality_score(
        df
    )

    print(
        "Quality Score calculated."
    )

    # ========================================================
    # STEP 14
    # GROWTH
    # ========================================================

    print_header(
        "CALCULATING GROWTH SCORE"
    )

    df = calculate_growth_score(
        df
    )

    print(
        "Growth Score calculated."
    )

    # ========================================================
    # STEP 15
    # VALUATION
    # ========================================================

    print_header(
        "CALCULATING VALUATION SCORE"
    )

    df = calculate_valuation_score(
        df
    )

    print(
        "Valuation Score calculated."
    )

    # ========================================================
    # STEP 16
    # FUNDAMENTAL SCORE
    # ========================================================

    print_header(
        "CALCULATING FUNDAMENTAL SCORE"
    )

    df = calculate_fundamental_score(
        df
    )

    print(
        "Fundamental Score calculated."
    )

    # ========================================================
    # STEP 17
    # COMBINED SCORE
    # ========================================================

    print_header(
        "CALCULATING COMBINED RESEARCH SCORE"
    )

    df = calculate_combined_research_score(
        df
    )

    print(
        "Combined Research Score calculated."
    )

    # ========================================================
    # STEP 18
    # UNIVERSE SUMMARY
    # ========================================================

    display_universe_summary(
        df
    )

    # ========================================================
    # STEP 19
    # MARKET RANKINGS
    # ========================================================

    display_market_rankings(
        df
    )

    # ========================================================
    # STEP 20
    # FUNDAMENTAL RANKINGS
    # ========================================================

    display_fundamental_rankings(
        df
    )

    # ========================================================
    # STEP 21
    # COMBINED RANKINGS
    # ========================================================

    display_combined_rankings(
        df
    )

    # ========================================================
    # STEP 22
    # DETAILED
    # ========================================================

    display_top_detailed(
        df
    )

    # ========================================================
    # STEP 23
    # FACTOR LEADERS
    # ========================================================

    display_factor_leaders(
        df
    )

    # ========================================================
    # STEP 24
    # RESEARCH CANDIDATES
    # ========================================================

    display_research_candidates(
        df
    )

    # ========================================================
    # STEP 25
    # SCORE DISTRIBUTION
    # ========================================================

    display_score_distribution(
        df
    )

    # ========================================================
    # STEP 26
    # FINAL TIMESTAMP
    #
    # Capture the completed research runtime BEFORE Excel
    # generation so Dashboard receives the actual runtime.
    # ========================================================

    PROGRAM_END_TIME = datetime.now()

    PROGRAM_ELAPSED_SECONDS = (
        time.perf_counter()
        - performance_start
    )

    # ========================================================
    # STEP 27
    # EXCEL OUTPUT
    # ========================================================

    print_header(
        "EXCEL REPORT"
    )

    build_excel_dashboard(
        df=df,
        universe_count=len(symbols),
        valid_symbol_count=len(valid_symbols),
        output_file=EXCEL_OUTPUT_FILE,
    )

    # ========================================================
    # STEP 28
    # RESEARCH NOTES
    # ========================================================

    display_research_notes()

    # ========================================================
    # FINAL SUMMARY
    # ========================================================

    display_final_summary(
        universe_count=len(symbols),
        valid_symbol_count=len(valid_symbols),
        df=df,
    )

    # ========================================================
    # RUNTIME
    # ========================================================

    print_header(
        "PROGRAM RUNTIME"
    )

    print(
        f"Start timestamp : "
        f"{PROGRAM_START_TIME.strftime('%Y-%m-%d %H:%M:%S')}"
    )

    print(
        f"End timestamp   : "
        f"{PROGRAM_END_TIME.strftime('%Y-%m-%d %H:%M:%S')}"
    )

    print(
        f"Elapsed seconds : "
        f"{PROGRAM_ELAPSED_SECONDS:.2f}"
    )

    elapsed_hours = int(
        PROGRAM_ELAPSED_SECONDS // 3600
    )

    elapsed_minutes = int(
        (
            PROGRAM_ELAPSED_SECONDS % 3600
        ) // 60
    )

    elapsed_seconds = (
        PROGRAM_ELAPSED_SECONDS
        % 60
    )

    print(
        f"Total runtime   : "
        f"{elapsed_hours:02d}:"
        f"{elapsed_minutes:02d}:"
        f"{elapsed_seconds:05.2f}"
    )

    print()

    print(
        f"Excel report    : "
        f"{EXCEL_OUTPUT_FILE}"
    )

    print()

    print(
        "=" * 80
    )

    print(
        "FUNDAMENTALALPHAFORGE RUN FINISHED SUCCESSFULLY"
    )

    print(
        "=" * 80
    )


# ============================================================
# PROGRAM ENTRY POINT
# ============================================================

if __name__ == "__main__":

    try:

        main()

    except KeyboardInterrupt:

        print()

        print(
            "Research run interrupted by user."
        )

        sys.exit(1)

    except Exception as error:

        PROGRAM_END_TIME = datetime.now()

        print()

        print("=" * 80)

        print(
            "RESEARCH RUN FAILED"
        )

        print("=" * 80)

        print()

        print(
            f"Error: {error}"
        )

        print()

        sys.exit(1)